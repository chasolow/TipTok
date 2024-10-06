import streamlit as st
import requests
from io import BytesIO
from PIL import Image
import pandas as pd
from openpyxl import load_workbook, Workbook
import os

# Подключаем стили и пр.
st.markdown("""
    <style>
        /* Текущие стили */
    </style>
""", unsafe_allow_html=True)

# URL изображения
image_url = "https://i.postimg.cc/vZmHCG8k/big.png"
response = requests.get(image_url)
image = Image.open(BytesIO(response.content))
st.image(image)

st.markdown("<h1>Расчет стоимости услуг</h1>", unsafe_allow_html=True)

power_question = st.radio("Есть ли на объекте существующая мощность?", ['⚡️ Да', '❌ Нет'], index=0)

if power_question == '⚡️ Да':
    P = st.number_input("Введите суммарную мощность (P, кВт):", min_value=0, max_value=500000, step=1, value=None)
    Pdop = st.number_input("Введите доп. мощность (Pдоп, кВт):", min_value=0, max_value=500000, step=1, value=None)
else:
    P = st.number_input("Введите суммарную мощность (P, кВт):", min_value=0, max_value=500000, step=1, value=None)
    Pdop = None

voltage_classes = {
    "До 1000 В": 1,
    "3 - 35 кВ": 1.1,
    "110 кВ": 1.2,
    "220 кВ": 1.3
}
Ku = st.radio("Класс напряжения:", list(voltage_classes.keys()))
Ku_value = voltage_classes[Ku]

Ktg = st.radio("Есть ли компенсация реактивной мощности?", ['✅ Есть', '❌ Нет'])
schemes = st.radio("Есть ли схемы питания?", ['✅ Есть', '❌ Нет'])

if schemes == '✅ Есть':
    X = st.number_input("Количество участков ЛЭП:", min_value=0, max_value=5000, step=1, value=None)
    Y = st.number_input("Количество электроприемников:", min_value=0, max_value=5000, step=1, value=None)
else:
    Y = st.number_input("Количество электроприемников:", min_value=0, max_value=5000, step=1, value=None)
    X = Y * 1.05

Kc = st.radio("Требуется ли сопровождение согласования?", ['📝 Требуется', '❌ Не требуется'])

# Функция для записи в Excel с логами
def save_to_excel(data):
    file_path = r'C:\Users\wanss\OneDrive\Рабочий стол\TipTok\Статистика\Калькулятор.xlsx'
    
    # Проверяем, существует ли файл
    if not os.path.exists(file_path):
        st.warning(f"Файл не найден: {file_path}, создаем новый.")
        workbook = Workbook()  # Создаем новую книгу
        sheet = workbook.active
        # Записываем заголовки в новую книгу
        sheet.append(['№', 'Тип услуги', 'P', 'Pдоп', 'U', 'КРМ', 'Схемы', 'Участки', 'ЭП', 'Согласование', 'Стоимость'])
        workbook.save(file_path)

    try:
        # Загружаем существующий файл Excel
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book

        # Читаем текущие данные из файла
        df_existing = pd.read_excel(file_path)

        # Определяем следующий порядковый номер
        next_index = len(df_existing) + 1

        # Добавляем новые данные в DataFrame
        df_new = pd.DataFrame(data)
        df_new.insert(0, '№', range(next_index, next_index + len(df_new)))  # Вставляем столбец с порядковыми номерами
        df_result = pd.concat([df_existing, df_new], ignore_index=True)

        # Сохраняем в файл
        df_result.to_excel(writer, index=False)
        writer.save()
        st.success("Данные успешно сохранены в Excel!")

    except Exception as e:
        st.error(f"Ошибка при записи в Excel: {e}")

# Кнопка расчета
if st.button('РАСЧЁТ'):
    try:
        if P <= 0:
            st.error("Параметры должны быть больше нуля")
        else:
            Kp = (0.8533 * Pdop ** 0.0599 + (0.8533 * P ** 0.0599 - 0.8533 * Pdop ** 0.0599) * Pdop / P) if Pdop else 0.8533 * P ** 0.0599
            Ktg_value = 1.1 if Ktg == '✅ Есть' else 1
            Kc_value = 1.05 if Kc == '📝 Требуется' else 1

            if schemes == '✅ Есть':
                Gx = 1892.9 * X ** -0.544
                Gy = 379.89 * Y ** -0.271
                Gz = 0  # Gz не рассчитываем
            else:
                Gx = 1892.9 * X ** -0.544
                Gy = 379.89 * Y ** -0.271
                Gz = 966.81 * 2 * X ** -0.424

            cost = round(((X * Gx + Y * Gy) * Kp * Ku_value * Ktg_value * Kc_value + X * Gz) / 100) * 100
            st.markdown(f"<h1 style='color: #F4B03F;'>Общая стоимость: {cost} рублей</h1>", unsafe_allow_html=True)

            # Данные для записи в Excel
            data = {
                'Тип услуги': ['КЭЭ'],
                'P': [P],
                'Pдоп': [Pdop],
                'U': [Ku],
                'КРМ': [Ktg],
                'Схемы': [schemes],
                'Участки': [X],
                'ЭП': [Y],
                'Согласование': [Kc],
                'Стоимость': [cost]
            }

            # Сохранение данных
            save_to_excel(data)

    except ZeroDivisionError:
        st.error("Ошибка: деление на ноль!")
    except Exception as e:
        st.error(f"Ошибка: {e}")
